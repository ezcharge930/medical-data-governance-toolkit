import openpyxl
import yaml

'''
excel转换为yaml的思路
1、读取excel文件
2、读取文件中的数据源定义
3、去掉标题行,按行读取每一行数据
4、将每一行数据转换为字典,添加到列表里面
5、读取文件中的字段映射
6、也是去掉标题行,按行读取每一行数据
7、将每一行数据同样转成为字典,添加到列表里面
8、将数据源定义和字段映射组成一个大字典
9、映射为yaml格式,写入到yaml文件里面.
'''

def excel_to_yaml(excel_path: str, yaml_path: str) -> None:
    ''' 读取excel文件'''
    wb = openpyxl.load_workbook(excel_path)
    
    # 读取数据源定义
    source_sheet = wb['数据源定义']
    source:list[dict] = []
    for row in source_sheet.iter_rows(min_row= 2, values_only= True):
        if row[0]:
            source.append({
                'table_name': row[0],
                'table_name_en': row[1],
                'owner': row[2]
            })
            
    mapping_sheet = wb['字段映射']
    fields: dict[str, dict] = {}
    special_rules: dict[str, list[str]] = {}
    for row in mapping_sheet.iter_rows(min_row= 2, values_only= True):
        if row[0] is None:
            continue
        
        rule_mark = str(row[7]).strip() if row[7] else None
        
        target_field: str = str(row[0])
        
        # 若目标变量第一次出现，初始化
        if target_field not in fields:
            fields[target_field] = {
                'description': row[1], # 描述
                'data_type': row[2],   # 数据类型
                'is_must': row[3],    # 是否必填
                'source_table': [],
                'special_rule': rule_mark # 特殊规则
            }
            if rule_mark:
                if rule_mark not in special_rules:
                    special_rules[rule_mark] = []
                special_rules[rule_mark].append(target_field)
        else: # 若目标变量已存在，检查特殊规则是否一致
            # todo 处理思路：
            ''' step1: 先判断rule_mark是否为空,并且目前的特殊规则与原先是否一致
                step2: 若不一致,将原来的old_rule取出来进行判断
                step3: old_rule不为空,old_rule在special_rules里面,对应的目标变量在old_rule下面
                step4: 在old_rule下面移除这个目标变量
                step5: 若恰好这个变量是最后一个变量,移除以后,删除这个old_rule
            '''
            if (rule_mark) and (fields[target_field].get('special_rule') != rule_mark):
                # 1.rule_mark不为空
                # 2.且与已有的规则不一致
                # 移除旧的规则标记
                old_rule = fields[target_field].get('special_rule') # 获取旧的规则标记
                if (old_rule) and (old_rule in special_rules) and (target_field in special_rules[rule_mark]):
                    # 1.旧的规则标记存在
                    # 2.且在special_rules中
                    # 3.且目标变量在该规则下
                    special_rules[old_rule].remove(target_field) # 从旧规则中移除目标变量
                    if not special_rules[old_rule]: # 若旧规则下没有变量了，删除该规则
                        del special_rules[old_rule]

                # 添加新的规则标记
                # todo 处理思路
                # step1: 将fields中的old_rule更新为rule_mark
                # step2: 判断rule_mark是否在special_rules中
                # step3: 若不在,则初始化一个空列表
                # step4: 判断目标变量是否在rule_mark下
                # step5: 若不在,则添加进去
                fields[target_field]['special_rule'] = rule_mark
                if rule_mark not in special_rules:
                    special_rules[rule_mark] = []
                if target_field not in special_rules[rule_mark]:
                    special_rules[rule_mark].append(target_field)

            
                
        source_table = row[4]
        source_field = row[5]
        priority = row[6] if row[6] is not None else 1
        
        fields[target_field]['source_table'].append({
            'source_table': source_table,
            'source_field': source_field,
            'priority': priority,
        })
               
            
    config = {
        'source': source,
        'fields': fields,
        'special_rules': special_rules
    }
    
    # 写入yaml文件
    with open(yaml_path, 'w', encoding= 'utf-8') as file:
        yaml.dump(config, file, allow_unicode= True, sort_keys= False)


# if __name__ == '__main__':
#     excel_to_yaml('test/DM_config.xlsx', 'test/DM_config.yaml')
#     pass